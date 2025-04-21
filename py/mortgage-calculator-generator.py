import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import datetime
import os

def create_mortgage_calculator():
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create Calculator sheet (main sheet)
    calculator_sheet = wb.active
    calculator_sheet.title = "Calculator"
    
    # Create other sheets
    amortization_sheet = wb.create_sheet("Amortization")
    comparison_sheet = wb.create_sheet("Loan Comparison")
    affordability_sheet = wb.create_sheet("Affordability")
    
    # Set up Calculator sheet
    setup_calculator_sheet(calculator_sheet)
    
    # Set up Amortization sheet
    setup_amortization_sheet(amortization_sheet)
    
    # Set up Loan Comparison sheet
    setup_comparison_sheet(comparison_sheet)
    
    # Set up Affordability sheet
    setup_affordability_sheet(affordability_sheet)
    
    # Get output directory from environment, default to current directory
    output_dir = os.environ.get('XL_OUTPUT_DIR', '.')
    
    # Save the workbook
    wb.save(os.path.join(output_dir, "Mortgage_Calculator.xlsx"))
    print("Mortgage calculator Excel file created successfully.")

def setup_calculator_sheet(sheet):
    # Set up header
    sheet.merge_cells('A1:G1')
    sheet['A1'] = "REAL ESTATE MORTGAGE CALCULATOR"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    # Create sections
    sheet['A3'] = "LOAN INFORMATION"
    sheet['A3'].font = Font(bold=True)
    sheet['A13'] = "PROPERTY TAXES & INSURANCE"
    sheet['A13'].font = Font(bold=True)
    sheet['A19'] = "CLOSING COSTS & FEES"
    sheet['A19'].font = Font(bold=True)
    sheet['A25'] = "PAYMENT SUMMARY"
    sheet['A25'].font = Font(bold=True)

    # Loan Information Section
    sheet['A4'] = "Purchase Price ($):"
    sheet['A5'] = "Down Payment ($):"
    sheet['A6'] = "Down Payment (%):"
    sheet['A7'] = "Loan Amount ($):"
    sheet['A8'] = "Interest Rate (%):"
    sheet['A9'] = "Loan Term (years):"
    sheet['A10'] = "Loan Start Date:"
    sheet['A11'] = "Payment Type:"

    # Set up input fields
    sheet['C4'] = 300000
    sheet['C5'] = 60000
    sheet['C6'] = "=C5/C4"
    sheet['C7'] = "=C4-C5"
    sheet['C8'] = 0.0575  # Store as decimal (5.75%)
    sheet['C9'] = 30
    sheet['C10'] = datetime.datetime.now().strftime("%m/%d/%Y")
    sheet['C11'] = "Standard"

    # Format cells
    for cell in ['C4', 'C5', 'C7']:
        sheet[cell].number_format = '$#,##0.00'
    sheet['C6'].number_format = '0.00%'
    sheet['C8'].number_format = '0.00%'
    sheet['C9'].number_format = '0'
    sheet['C10'].number_format = 'mm/dd/yyyy'

    # Create loan type dropdown
    dv = DataValidation(type="list", formula1='"Standard,Balloon,Interest Only"')
    sheet.add_data_validation(dv)
    dv.add(sheet['C11'])

    # Property Taxes & Insurance Section
    sheet['A14'] = "Annual Property Tax ($):"
    sheet['A15'] = "Annual Property Tax Rate (%):"
    sheet['A16'] = "Annual Homeowners Insurance ($):"
    sheet['A17'] = "Monthly PMI (%):"

    sheet['C14'] = 3000
    sheet['C15'] = "=C14/C4"
    sheet['C16'] = 1200
    sheet['C17'] = 0.005  # 0.5% as decimal

    sheet['C14'].number_format = '$#,##0.00'
    sheet['C16'].number_format = '$#,##0.00'
    sheet['C15'].number_format = '0.00%'
    sheet['C17'].number_format = '0.00%'

    # Conditional PMI calculation
    sheet['E17'] = "=IF(C5/C4<0.2,C7*C17/12,0)"
    sheet['E17'].number_format = '$#,##0.00'
    sheet['D17'] = "Monthly PMI Amount:"

    # Closing Costs & Fees Section
    sheet['A20'] = "Loan Origination Fee (%):"
    sheet['A21'] = "Other Closing Costs ($):"
    sheet['A22'] = "Total Closing Costs ($):"

    sheet['C20'] = 0.01  # 1% as decimal
    sheet['C21'] = 2500
    sheet['C22'] = "=(C7*C20)+C21"  # Removed /100 since C20 is already decimal

    sheet['C20'].number_format = '0.00%'
    sheet['C21'].number_format = '$#,##0.00'
    sheet['C22'].number_format = '$#,##0.00'

    # Payment Summary Section
    sheet['A26'] = "Principal & Interest Payment:"
    sheet['A27'] = "Monthly Property Tax:"
    sheet['A28'] = "Monthly Insurance:"
    sheet['A29'] = "Monthly PMI:"
    sheet['A30'] = "Total Monthly Payment:"
    sheet['A31'] = "Balloon Payment (if applicable):"

    # Calculate P&I payment based on loan type - no need to convert percentages
    sheet['C26'] = "=IF(C11=\"Standard\",PMT(C8/12,C9*12,-C7),IF(C11=\"Interest Only\",C7*C8/12,PMT(C8/12,C9*12,-C7,-(C7*0.7))))"
    sheet['C27'] = "=C14/12"
    sheet['C28'] = "=C16/12"
    sheet['C29'] = "=E17"
    sheet['C30'] = "=SUM(C26:C29)"
    sheet['C31'] = "=IF(C11=\"Balloon\",C7*0.7,0)"

    for cell in ['C26', 'C27', 'C28', 'C29', 'C30', 'C31']:
        sheet[cell].number_format = '$#,##0.00'

    # Add balloon payment explanation
    sheet['E31'] = "=IF(C11=\"Balloon\",\"(Due at end of term)\",\"\")"

def setup_amortization_sheet(sheet):
    # Set up header
    sheet.merge_cells('A1:G1')
    sheet['A1'] = "AMORTIZATION SCHEDULE"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')

    # Create column headers
    headers = {
        'A3': 'Payment #',
        'B3': 'Payment Date',
        'C3': 'Beginning Balance',
        'D3': 'Payment Amount',
        'E3': 'Principal',
        'F3': 'Interest',
        'G3': 'Ending Balance',
        'H3': 'Cumulative Interest',
        'I3': 'Balloon Payment'
    }
    
    for cell, value in headers.items():
        sheet[cell] = value
        sheet[cell].font = Font(bold=True)

    # Format columns - use 361 to accommodate a 30-year loan (360 payments) plus header row
    for row in range(4, 365):
        sheet[f'B{row}'].number_format = 'mm/dd/yyyy'
        for col in 'CDEFGHI':
            sheet[f'{col}{row}'].number_format = '$#,##0.00'

    # Set up formulas for amortization calculation
    sheet['A4'] = 1
    sheet['B4'] = "=EDATE(Calculator!C10,1)"  # First payment is one month after loan start
    sheet['C4'] = "=Calculator!C7"
    sheet['D4'] = "=IF(Calculator!C11=\"Standard\",ABS(Calculator!C26),IF(Calculator!C11=\"Interest Only\",C4*Calculator!C8/12,ABS(Calculator!C26)))"
    sheet['E4'] = "=IF(Calculator!C11=\"Interest Only\",0,D4-F4)"
    sheet['F4'] = "=C4*Calculator!C8/12"
    sheet['G4'] = "=C4-E4"
    sheet['H4'] = "=F4"
    sheet['I4'] = "=IF(Calculator!C11=\"Balloon\",IF(A4=Calculator!C9*12,Calculator!C31,0),0)"

    # Add formulas for subsequent rows
    for row in range(5, 365):
        sheet[f'A{row}'] = f"=A{row-1}+1"
        sheet[f'B{row}'] = f"=EDATE(B{row-1},1)"
        sheet[f'C{row}'] = f"=G{row-1}"
        sheet[f'D{row}'] = f"=IF(Calculator!C11=\"Standard\",ABS(Calculator!C26),IF(Calculator!C11=\"Interest Only\",C{row}*Calculator!C8/12,ABS(Calculator!C26)))"
        sheet[f'E{row}'] = f"=IF(Calculator!C11=\"Interest Only\",0,D{row}-F{row})"
        sheet[f'F{row}'] = f"=C{row}*Calculator!C8/12"
        sheet[f'G{row}'] = f"=C{row}-E{row}"
        sheet[f'H{row}'] = f"=H{row-1}+F{row}"
        sheet[f'I{row}'] = f"=IF(Calculator!C11=\"Balloon\",IF(A{row}=Calculator!C9*12,Calculator!C31,0),0)"

def setup_comparison_sheet(sheet):
    # Set up header
    sheet.merge_cells('A1:G1')
    sheet['A1'] = "LOAN COMPARISON CALCULATOR"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')

    # Create comparison table
    sheet.merge_cells('A3:G3')
    sheet['A3'] = "COMPARE DIFFERENT LOAN OPTIONS"
    sheet['A3'].font = Font(bold=True)

    # Create table headers
    headers = {
        'A5': 'Loan Option',
        'B5': 'Rate (%)',
        'C5': 'Term (Years)',
        'D5': 'Monthly P&I',
        'E5': 'Total Monthly Payment',
        'F5': 'Total Interest Paid',
        'G5': 'Total Cost'
    }
    
    for cell, value in headers.items():
        sheet[cell] = value
        sheet[cell].font = Font(bold=True)

    # Set up comparison rows
    sheet['A6'] = "Option 1 (Current)"
    sheet['A7'] = "Option 2"
    sheet['A8'] = "Option 3"

    # Link first option to main calculator
    sheet['B6'] = "=Calculator!C8"
    sheet['C6'] = "=Calculator!C9"
    sheet['D6'] = "=ABS(Calculator!C26)"
    sheet['E6'] = "=Calculator!C30"
    sheet['F6'] = "=IF(Calculator!C11=\"Standard\",INDIRECT(\"Amortization!H\"&Calculator!C9*12+3),Calculator!C8/12*Calculator!C7*Calculator!C9)"
    sheet['G6'] = "=F6+Calculator!C7+Calculator!C22"

    # Set up option 2 with example values (store as decimals)
    sheet['B7'] = 0.06  # 6%
    sheet['C7'] = 15
    sheet['D7'] = "=PMT(B7/12,C7*12,-Calculator!C7)"
    sheet['E7'] = "=D7+Calculator!C27+Calculator!C28+Calculator!C29"
    sheet['F7'] = "=D7*C7*12-Calculator!C7"
    sheet['G7'] = "=F7+Calculator!C7+Calculator!C22"

    # Set up option 3 with example values (store as decimals)
    sheet['B8'] = 0.055  # 5.5%
    sheet['C8'] = 30
    sheet['D8'] = "=PMT(B8/12,C8*12,-Calculator!C7)"
    sheet['E8'] = "=D8+Calculator!C27+Calculator!C28+Calculator!C29"
    sheet['F8'] = "=D8*C8*12-Calculator!C7"
    sheet['G8'] = "=F8+Calculator!C7+Calculator!C22"

    # Format cells
    for row in range(6, 9):
        sheet[f'B{row}'].number_format = '0.00%'
        sheet[f'C{row}'].number_format = '0'
        for col in 'DEFG':
            sheet[f'{col}{row}'].number_format = '$#,##0.00'

def setup_affordability_sheet(sheet):
    # Set up header
    sheet.merge_cells('A1:G1')
    sheet['A1'] = "AFFORDABILITY CALCULATOR"
    sheet['A1'].font = Font(size=16, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center')

    # Create income and expense inputs
    sheet['A3'] = "INCOME & EXPENSE INFORMATION"
    sheet['A3'].font = Font(bold=True)

    sheet['A5'] = "Gross Annual Income ($):"
    sheet['A6'] = "Monthly Debt Payments ($):"
    sheet['A7'] = "Desired Monthly Payment ($):"
    sheet['A8'] = "Maximum DTI Ratio (%):"

    sheet['C5'] = 85000
    sheet['C6'] = 500
    sheet['C7'] = "=C5/12*0.28"
    sheet['C8'] = 0.36  # 36% as decimal

    sheet['C5'].number_format = '$#,##0.00'
    sheet['C6'].number_format = '$#,##0.00'
    sheet['C7'].number_format = '$#,##0.00'
    sheet['C8'].number_format = '0.00%'

    # Create affordability calculator
    sheet['A10'] = "AFFORDABILITY RESULTS"
    sheet['A10'].font = Font(bold=True)

    sheet['A12'] = "Interest Rate (%):"
    sheet['A13'] = "Loan Term (years):"
    sheet['A14'] = "Property Tax Rate (%):"
    sheet['A15'] = "Annual Insurance Rate (%):"
    sheet['A16'] = "Down Payment (%):"

    sheet['C12'] = "=Calculator!C8"
    sheet['C13'] = "=Calculator!C9"
    sheet['C14'] = "=Calculator!C15"
    sheet['C15'] = "=Calculator!C16/Calculator!C4"
    sheet['C16'] = "=Calculator!C6"

    sheet['C12'].number_format = '0.00%'
    sheet['C13'].number_format = '0'
    sheet['C14'].number_format = '0.00%'
    sheet['C15'].number_format = '0.00%'
    sheet['C16'].number_format = '0.00%'

    sheet['A18'] = "Maximum Affordable Loan:"
    sheet['A19'] = "Maximum Home Price:"
    sheet['A20'] = "Down Payment Amount:"
    sheet['A21'] = "Monthly Principal & Interest:"
    sheet['A22'] = "Monthly Property Taxes:"
    sheet['A23'] = "Monthly Insurance:"
    sheet['A24'] = "Total Monthly Payment:"

    # Calculate maximum affordable home price based on income and expenses
    # Fixed formulas to work with percentage values stored as decimals
    sheet['C18'] = "=PV(C12/12,C13*12,-MAX(0,(C5/12*C8)-C6))"
    sheet['C19'] = "=C18/(1-C16)"
    sheet['C20'] = "=C19*C16"
    sheet['C21'] = "=PMT(C12/12,C13*12,-C18)"
    sheet['C22'] = "=C19*C14/12"
    sheet['C23'] = "=C19*C15/12"
    sheet['C24'] = "=SUM(ABS(C21),C22,C23)"

    for row in range(18, 25):
        sheet[f'C{row}'].number_format = '$#,##0.00'

if __name__ == "__main__":
    create_mortgage_calculator()
