import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import PieChart, LineChart, Reference
from datetime import datetime, timedelta
import random
import math
import os
import traceback

def generate_real_estate_budget(filename="Real_Estate_Development_Budget.xlsx", 
                               project_name="Your Project Name",
                               project_address="Your Project Address",
                               project_size=100000,  # SF
                               project_type="Commercial Development",
                               forecast_periods=24,
                               forecast_type="Monthly"):
    """Generate a comprehensive real estate development budget Excel workbook."""
    print(f"Generating real estate development budget for: {project_name}")
    
    try:
        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        
        # Define basic structure for budget
        budget_structure = {
            "hardCosts": [
                {"category": "Site Work", "items": ["Site Preparation", "Excavation", "Landscaping"]},
                {"category": "Structure", "items": ["Framing", "Concrete", "Roofing"]},
                {"category": "Mechanical", "items": ["HVAC", "Plumbing", "Electrical"]},
                {"category": "Finishes", "items": ["Drywall", "Flooring", "Paint"]},
                {"category": "Contingency", "items": ["Construction Contingency"]}
            ],
            "softCosts": [
                {"category": "Professional Fees", "items": ["Architecture", "Engineering", "Legal"]},
                {"category": "Permits & Fees", "items": ["Building Permits", "Impact Fees"]},
                {"category": "Financing", "items": ["Loan Fees", "Interest"]},
                {"category": "Marketing", "items": ["Marketing Materials", "Promotions"]},
                {"category": "Contingency", "items": ["Soft Cost Contingency"]}
            ],
            "other": [
                {"category": "Land", "items": ["Land Acquisition", "Closing Costs"]},
                {"category": "Taxes & Insurance", "items": ["Property Taxes", "Insurance"]}
            ]
        }

        # Generate budget data
        budget_data = {}
        section_totals = {"hardCosts": 0, "softCosts": 0, "other": 0}
        category_totals = {}
        line_items_data = []
        
        # Values to control relative size of each section
        section_multipliers = {"hardCosts": 5, "softCosts": 3, "other": 2}
        
        # Process each section
        for section, categories in budget_structure.items():
            section_multiplier = section_multipliers[section]
            
            for category in categories:
                category_name = category["category"]
                category_total = 0
                
                for item in category["items"]:
                    # Generate a realistic amount
                    base_amount = random.randint(5, 50) * 10000 * section_multiplier
                    
                    # Adjust special items
                    if item == "Land Acquisition":
                        base_amount = random.randint(30, 50) * 100000
                    elif "Contingency" in item:
                        # Set contingency as % of costs so far
                        if section == "hardCosts":
                            base_amount = int(section_totals["hardCosts"] * 0.05)
                        elif section == "softCosts":
                            base_amount = int(section_totals["softCosts"] * 0.05)
                    
                    budget_data[f"{section}_{category_name}_{item}"] = base_amount
                    category_total += base_amount
                    
                    # Store line item data
                    line_items_data.append({
                        "section": section,
                        "section_name": section.replace("hardCosts", "HARD COSTS")
                                              .replace("softCosts", "SOFT COSTS")
                                              .replace("other", "OTHER COSTS"),
                        "category": category_name,
                        "item": item,
                        "amount": base_amount
                    })
                
                category_totals[f"{section}_{category_name}"] = category_total
                section_totals[section] += category_total
        
        total_budget = sum(section_totals.values())
        
        # Set up sheets
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        summary_sheet = wb.create_sheet("Summary", 0)
        detailed_sheet = wb.create_sheet("Detailed Budget", 1)
        forecast_sheet = wb.create_sheet("Forecast", 2)
        
        # Basic setup for summary sheet
        summary_sheet['A1'] = "REAL ESTATE DEVELOPMENT BUDGET"
        summary_sheet['A1'].font = Font(bold=True, size=16)
        summary_sheet.merge_cells('A1:C1')
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Project Information
        row = 3
        summary_sheet['A3'] = "Project Name:"
        summary_sheet['B3'] = project_name
        summary_sheet['A4'] = "Project Address:"
        summary_sheet['B4'] = project_address
        summary_sheet['A5'] = "Project Size:"
        summary_sheet['B5'] = f"{project_size:,} SF"
        summary_sheet['A6'] = "Date Created:"
        summary_sheet['B6'] = datetime.now().strftime("%m/%d/%Y")
        
        # Budget Summary
        summary_sheet['A8'] = "BUDGET SUMMARY"
        summary_sheet['A8'].font = Font(bold=True)
        
        summary_sheet['A10'] = "Category"
        summary_sheet['B10'] = "Amount"
        summary_sheet['C10'] = "% of Total"
        
        # Hard Costs
        summary_sheet['A11'] = "Hard Costs"
        summary_sheet['B11'] = section_totals["hardCosts"]
        summary_sheet['B11'].number_format = '"$"#,##0'
        summary_sheet['C11'] = section_totals["hardCosts"] / total_budget
        summary_sheet['C11'].number_format = '0.0%'
        
        # Soft Costs
        summary_sheet['A12'] = "Soft Costs"
        summary_sheet['B12'] = section_totals["softCosts"]
        summary_sheet['B12'].number_format = '"$"#,##0'
        summary_sheet['C12'] = section_totals["softCosts"] / total_budget
        summary_sheet['C12'].number_format = '0.0%'
        
        # Other Costs
        summary_sheet['A13'] = "Other Costs"
        summary_sheet['B13'] = section_totals["other"]
        summary_sheet['B13'].number_format = '"$"#,##0'
        summary_sheet['C13'] = section_totals["other"] / total_budget
        summary_sheet['C13'].number_format = '0.0%'
        
        # Total
        summary_sheet['A14'] = "TOTAL"
        summary_sheet['A14'].font = Font(bold=True)
        summary_sheet['B14'] = total_budget
        summary_sheet['B14'].number_format = '"$"#,##0'
        summary_sheet['C14'] = 1.0
        summary_sheet['C14'].number_format = '0.0%'
        
        # Add simple pie chart (in a try block in case of chart issues)
        try:
            chart = PieChart()
            chart.title = "Budget Breakdown"
            
            data = Reference(summary_sheet, min_col=2, min_row=11, max_row=13)
            cats = Reference(summary_sheet, min_col=1, min_row=11, max_row=13)
            
            chart.add_data(data)
            chart.set_categories(cats)
            
            chart.height = 10
            chart.width = 10
            
            summary_sheet.add_chart(chart, 'A16')
            print("Added pie chart to summary sheet")
        except Exception as e:
            print(f"Warning: Could not create pie chart, continuing without it: {str(e)}")
        
        # Basic setup for detailed sheet
        detailed_sheet['A1'] = "DETAILED DEVELOPMENT BUDGET"
        detailed_sheet['A1'].font = Font(bold=True, size=16)
        detailed_sheet.merge_cells('A1:E1')
        detailed_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Column Headers
        detailed_sheet['A3'] = "Category"
        detailed_sheet['B3'] = "Line Item"
        detailed_sheet['C3'] = "Budget Amount"
        detailed_sheet['D3'] = "Cost per SF"
        detailed_sheet['E3'] = "% of Total"
        
        for cell in ['A3', 'B3', 'C3', 'D3', 'E3']:
            detailed_sheet[cell].font = Font(bold=True)
        
        # Add line items (simplified)
        row = 4
        for item in line_items_data:
            row += 1
            detailed_sheet[f'A{row}'] = item["section_name"] + ": " + item["category"]
            detailed_sheet[f'B{row}'] = item["item"]
            detailed_sheet[f'C{row}'] = item["amount"]
            detailed_sheet[f'C{row}'].number_format = '"$"#,##0'
            detailed_sheet[f'D{row}'] = item["amount"] / project_size
            detailed_sheet[f'D{row}'].number_format = '"$"#,##0.00'
            detailed_sheet[f'E{row}'] = item["amount"] / total_budget
            detailed_sheet[f'E{row}'].number_format = '0.0%'
        
        # Add total row
        row += 2
        detailed_sheet[f'A{row}'] = "TOTAL PROJECT BUDGET"
        detailed_sheet[f'A{row}'].font = Font(bold=True)
        detailed_sheet[f'C{row}'] = total_budget
        detailed_sheet[f'C{row}'].number_format = '"$"#,##0'
        detailed_sheet[f'E{row}'] = 1.0
        detailed_sheet[f'E{row}'].number_format = '0.0%'
        
        # Basic forecast sheet
        forecast_sheet['A1'] = "FORECAST - TO BE IMPLEMENTED"
        forecast_sheet['A1'].font = Font(bold=True, size=16)
        
        # Save the workbook
        wb.save(filename)
        print(f"Excel workbook created successfully at: {filename}")
        return filename
    
    except Exception as e:
        print(f"ERROR: Failed to generate budget: {str(e)}")
        traceback.print_exc()
        return None

if __name__ == "__main__":
    try:
        # Output directory handling with robust error checking
        output_dir = os.environ.get('XL_OUTPUT_DIR', '.')
        
        # Ensure directory exists
        if not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir)
                print(f"Created output directory: {output_dir}")
            except PermissionError:
                print("WARNING: Cannot create output directory due to permission error.")
                print("Using current directory instead.")
                output_dir = '.'
        
        # Make sure we have write access to the directory
        test_file = os.path.join(output_dir, "test_write.txt")
        try:
            with open(test_file, 'w') as f:
                f.write("test")
            os.remove(test_file)
        except (PermissionError, IOError):
            print("WARNING: Cannot write to output directory due to permission error.")
            print("Using current directory instead.")
            output_dir = '.'
        
        # Full path for output file with timestamp to avoid overwriting
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"Real_Estate_Development_Budget_{timestamp}.xlsx")
        
        # Generate the budget
        result = generate_real_estate_budget(filename=output_file)
        
        if result:
            print(f"SUCCESS: Budget file created at: {result}")
        else:
            print("ERROR: Budget file generation failed.")
    
    except Exception as e:
        print(f"FATAL ERROR: {str(e)}")
        traceback.print_exc()